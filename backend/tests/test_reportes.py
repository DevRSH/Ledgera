import pytest
import io
import openpyxl
from app.services.reporte_service import reporte_service

@pytest.mark.asyncio
async def test_exportar_xlsx_valido():
    datos = [{"id": 1, "nombre": "Test"}, {"id": 2, "nombre": "Example"}]
    columnas = [
        {"header": "ID", "key": "id"},
        {"header": "Nombre", "key": "nombre"}
    ]
    
    xlsx_bytes = await reporte_service.exportar_xlsx(datos, columnas, "Test Report")
    
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0
    
    # Verify content with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "Test Report"
    assert ws.cell(row=1, column=1).value == "ID"
    assert ws.cell(row=2, column=2).value == "Test"
