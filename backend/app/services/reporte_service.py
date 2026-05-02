import io
from typing import List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from app.services.pdf_service import pdf_service

class ReporteService:
    async def exportar_xlsx(self, datos: List[dict], columnas: List[dict], titulo: str) -> bytes:
        """Helper to export any dataset to Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:31]
        
        # Header styling
        header_fill = PatternFill("solid", fgColor="059669")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, col in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Data
        for row_idx, row in enumerate(datos, 2):
            for col_idx, col in enumerate(columnas, 1):
                value = row.get(col["key"])
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def balance_mensual_pdf(self, context: dict) -> bytes:
        # This would use a specific template for balance
        # For now, let's assume we use the PDF service with a new template
        pass

reporte_service = ReporteService()
